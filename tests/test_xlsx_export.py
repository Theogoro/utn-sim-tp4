import datetime
import json

from openpyxl import load_workbook

from backend.models import SimulationModel
from backend.services import xlsx_export


def _make_sim():
    sim = SimulationModel(
        num_pcs=2, min_enrollment=5.0, max_enrollment=8.0, mean_arrival_time=2.0,
        min_service_time=5.0, max_service_time=8.0, min_maintenance_time=3.0,
        max_maintenance_time=10.0, mean_technician_return_time=60.0,
        technician_return_time_variation=3.0, student_wait_threshold=5,
        student_return_time=30.0, initial_maintenance_at_start=True, sim_days=1.0,
    )
    sim.id = 1
    sim.created_at = datetime.datetime(2026, 1, 1)
    sim.total_students_arrived = 0
    sim.total_new_students_arrived = 0
    sim.total_students_returned = 0
    sim.registrations_completed = 0
    sim.total_technician_visits = 0
    sim.pct_students_returned = 0.0
    sim.avg_waiting_time = 0.0
    sim.avg_technician_idle_time = 0.0
    sim.pc_utilization = json.dumps(
        [{"id": 1, "busy_time": 0, "maintenance_time": 0, "idle_time": 0}]
    )
    return sim


def test_build_workbook_returns_three_named_sheets():
    buf = xlsx_export.build_workbook(_make_sim(), iter([]), iter([]))
    wb = load_workbook(buf)
    assert wb.sheetnames == ["Resumen", "Vector de Estados", "Detalle Alumnos"]
    assert buf.getvalue()[:2] == b"PK"
