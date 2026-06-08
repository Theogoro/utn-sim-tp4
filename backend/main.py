import sys
import os

# Ensure the root directory is on the path so we can import 'simulation' and 'utils'
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any

from backend.database import engine, Base, get_db
from backend.models import SimulationModel, SimulationLineModel, SimulationStudentModel
from backend.schemas import SimulationParamsCreate, SimulationResponse, SimulationLineResponse, SimulationStudentResponse
from backend.db_logger import DatabaseLoggerHandler
from backend.repositories.simulation_repository import SimulationRepository

from simulation.params import SimulationParams
from simulation.simulation import Simulation

# Create the database tables
Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="UTN Enrollment Simulation API",
    description="Backend API for discrete event simulation of the UTN exam enrollment process.",
    version="1.0.0"
)

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # For local development we can allow all origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/api/simulations", response_model=SimulationResponse)
def run_simulation(params_in: SimulationParamsCreate, db: Session = Depends(get_db)):
    """
    Executes a discrete event simulation with the provided parameters,
    computes key enrollment metrics, and streams each state-vector row
    into the database within a single transaction.
    """
    import json
    try:
        # 1. Configure the simulation parameters object
        sim_params = SimulationParams(
            num_pcs=params_in.num_pcs,
            min_enrollment=params_in.min_enrollment * 60.0,
            max_enrollment=params_in.max_enrollment * 60.0,
            mean_arrival_time=params_in.mean_arrival_time * 60.0,
            min_service_time=params_in.min_enrollment * 60.0,
            max_service_time=params_in.max_enrollment * 60.0,
            min_maintenance_time=params_in.min_maintenance_time * 60.0,
            max_maintenance_time=params_in.max_maintenance_time * 60.0,
            mean_technician_return_time=params_in.mean_technician_return_time * 60.0,
            technician_return_time_variation=params_in.technician_return_time_variation * 60.0,
            student_wait_threshold=params_in.student_wait_threshold,
            student_return_time=params_in.student_return_time * 60.0,
            initial_maintenance_at_start=params_in.initial_maintenance_at_start,
        )
        max_simulation_time = params_in.sim_hours * 3600

        # 2. Pre-crear sim_model con ceros para obtener id antes de correr la simulación.
        #    Las líneas se streamean dentro de la misma transacción.
        sim_model = SimulationModel(
            num_pcs=params_in.num_pcs,
            min_enrollment=params_in.min_enrollment,
            max_enrollment=params_in.max_enrollment,
            mean_arrival_time=params_in.mean_arrival_time,
            min_service_time=params_in.min_service_time,
            max_service_time=params_in.max_service_time,
            min_maintenance_time=params_in.min_maintenance_time,
            max_maintenance_time=params_in.max_maintenance_time,
            mean_technician_return_time=params_in.mean_technician_return_time,
            technician_return_time_variation=params_in.technician_return_time_variation,
            student_wait_threshold=params_in.student_wait_threshold,
            student_return_time=params_in.student_return_time,
            initial_maintenance_at_start=params_in.initial_maintenance_at_start,
            sim_days=params_in.sim_hours / 24.0,
        )
        db.add(sim_model)
        db.flush()  # Asigna sim_model.id sin commitear todavía.

        # 3. Correr simulación con logger que streamea líneas y alumnos finalizados a la session.
        sim = Simulation(sim_params)
        logger = DatabaseLoggerHandler(sim.state, SimulationRepository(db), sim_model.id)
        sim.observers = [logger]
        final_state = sim.run(max_simulation_time)
        # Al cortar por horizonte pueden quedar alumnos activos; se guardan como estado final parcial.
        logger.flush_student_records(include_active=True)

        # 4. Calcular estadísticas finales y actualizar sim_model.
        stats = final_state.stats

        pct_students_returned = 0.0
        if stats.total_students_arrived > 0:
            pct_students_returned = (stats.total_students_returned / stats.total_students_arrived) * 100

        avg_waiting_time = 0.0
        if stats.students_queued_and_waited > 0:
            avg_waiting_time = stats.total_waiting_time / stats.students_queued_and_waited

        avg_technician_idle_time = 0.0
        if stats.total_technician_visits > 0:
            avg_technician_idle_time = stats.total_technician_idle_time / stats.total_technician_visits

        pc_util_list = []
        for pc in final_state.servers:
            idle_time = max(0.0, max_simulation_time - pc.busy_time - pc.maintenance_time)
            pc_util_list.append({
                "id": pc.id,
                "busy_time": pc.busy_time,
                "maintenance_time": pc.maintenance_time,
                "idle_time": idle_time,
            })

        sim_model.total_students_arrived = stats.total_students_arrived
        sim_model.total_new_students_arrived = stats.total_new_students_arrived
        sim_model.total_students_returned = stats.total_students_returned
        sim_model.registrations_completed = stats.registrations_completed
        sim_model.total_technician_visits = stats.total_technician_visits
        sim_model.pct_students_returned = pct_students_returned
        sim_model.avg_waiting_time = avg_waiting_time
        sim_model.avg_technician_idle_time = avg_technician_idle_time
        sim_model.pc_utilization = json.dumps(pc_util_list)

        # 5. Commit único: sim_model + todas las líneas en una sola transacción.
        db.commit()
        db.refresh(sim_model)
        return sim_model

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Simulation error: {str(e)}")


@app.get("/api/simulations", response_model=List[SimulationResponse])
def list_simulations(db: Session = Depends(get_db)):
    """
    Returns a list of all simulation summaries, ordered by creation date descending.
    Excludes the detailed vector lines to keep payload size lightweight.
    """
    return db.query(SimulationModel).order_by(SimulationModel.created_at.desc()).all()


@app.get("/api/simulations/{simulation_id}", response_model=SimulationResponse)
def get_simulation(simulation_id: int, db: Session = Depends(get_db)):
    """
    Returns details of a specific simulation by ID.
    """
    sim = db.query(SimulationModel).filter(SimulationModel.id == simulation_id).first()
    if not sim:
        raise HTTPException(status_code=404, detail="Simulation not found")
    return sim


@app.get("/api/simulations/{simulation_id}/lines")
def get_simulation_lines(
    simulation_id: int, 
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(50, ge=1, le=1000, description="Items per page"),
    db: Session = Depends(get_db)
):
    """
    Returns a paginated list of state vector lines for a given simulation.
    Crucial for front-end rendering performance when running long simulations.
    """
    # Check if simulation exists first
    sim = db.query(SimulationModel).filter(SimulationModel.id == simulation_id).first()
    if not sim:
        raise HTTPException(status_code=404, detail="Simulation not found")
        
    total_count = db.query(SimulationLineModel).filter(SimulationLineModel.simulation_id == simulation_id).count()
    
    offset = (page - 1) * limit
    items = db.query(SimulationLineModel)\
              .filter(SimulationLineModel.simulation_id == simulation_id)\
              .order_by(SimulationLineModel.line_index.asc())\
              .offset(offset)\
              .limit(limit)\
              .all()
              
    return {
        "total": total_count,
        "page": page,
        "limit": limit,
        "items": [SimulationLineResponse.model_validate(item) for item in items]
    }


@app.get("/api/simulations/{simulation_id}/students")
def get_simulation_students(
    simulation_id: int,
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(50, ge=1, le=1000, description="Items per page"),
    db: Session = Depends(get_db)
):
    """Detalle historico por alumno, incluidos los que ya no viven en memoria."""
    sim = db.query(SimulationModel).filter(SimulationModel.id == simulation_id).first()
    if not sim:
        raise HTTPException(status_code=404, detail="Simulation not found")

    query = db.query(SimulationStudentModel).filter(SimulationStudentModel.simulation_id == simulation_id)
    total_count = query.count()
    offset = (page - 1) * limit
    items = query.order_by(SimulationStudentModel.student_id.asc()).offset(offset).limit(limit).all()
    return {
        "total": total_count,
        "page": page,
        "limit": limit,
        "items": [SimulationStudentResponse.model_validate(item) for item in items],
    }


@app.get("/api/simulations/{simulation_id}/pc_stats")
def get_pc_utilization(simulation_id: int, db: Session = Depends(get_db)):
    """
    Computes and returns the state distribution (Idle, Busy, Maintenance)
    for each PC to render utilization charts.
    """
    import json
    sim = db.query(SimulationModel).filter(SimulationModel.id == simulation_id).first()
    if not sim:
        raise HTTPException(status_code=404, detail="Simulation not found")
        
    if sim.pc_utilization:
        return json.loads(sim.pc_utilization)
    return []


@app.get("/api/simulations/{simulation_id}/export")
def export_simulation_xlsx(simulation_id: int, db: Session = Depends(get_db)):
    """
    Exports a simulation (summary + full state vector trace) as an .xlsx file.
    """
    import io
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    sim = db.query(SimulationModel).filter(SimulationModel.id == simulation_id).first()
    if not sim:
        raise HTTPException(status_code=404, detail="Simulation not found")

    wb = Workbook()

    # --- Sheet 1: Resumen ---
    ws_sum = wb.active
    ws_sum.title = "Resumen"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    section_font = Font(bold=True, size=12, color="6366F1")

    ws_sum["A1"] = f"Simulación #{sim.id}"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Ejecutada: {sim.created_at}"

    sections = [
        ("Parámetros", [
            ("Cantidad de PCs", sim.num_pcs),
            ("Tiempo entre llegadas (min)", sim.mean_arrival_time),
            ("Inscripción mín (min)", sim.min_enrollment),
            ("Inscripción máx (min)", sim.max_enrollment),
            ("Mantenimiento mín (min)", sim.min_maintenance_time),
            ("Mantenimiento máx (min)", sim.max_maintenance_time),
            ("Frecuencia regreso técnico (min)", sim.mean_technician_return_time),
            ("Variación regreso técnico ± (min)", sim.technician_return_time_variation),
            ("Límite cola alumnos", sim.student_wait_threshold),
            ("Demora retorno alumno (min)", sim.student_return_time),
            ("Mantenimiento inicial en minuto 0", "Sí" if sim.initial_maintenance_at_start else "No"),
            ("Tiempo simulado (días)", sim.sim_days),
            ("Tiempo simulado (horas)", round(sim.sim_days * 24, 2)),
        ]),
        ("Métricas", [
            ("Alumnos arribados (totales)", sim.total_students_arrived),
            ("Alumnos nuevos", sim.total_new_students_arrived),
            ("Alumnos retirados (rechazados)", sim.total_students_returned),
            ("Inscripciones completadas", sim.registrations_completed),
            ("Visitas del técnico", sim.total_technician_visits),
            ("% intentos rechazados", round(sim.pct_students_returned, 4)),
            ("Espera promedio (seg)", round(sim.avg_waiting_time, 4)),
            ("Espera promedio (min)", round(sim.avg_waiting_time / 60.0, 4)),
            ("Ocio promedio técnico (seg)", round(sim.avg_technician_idle_time, 4)),
            ("Ocio promedio técnico (min)", round(sim.avg_technician_idle_time / 60.0, 4)),
        ]),
    ]

    row = 4
    for title, rows in sections:
        ws_sum.cell(row=row, column=1, value=title).font = section_font
        row += 1
        for label, val in rows:
            ws_sum.cell(row=row, column=1, value=label)
            ws_sum.cell(row=row, column=2, value=val)
            row += 1
        row += 1

    # PC utilization
    if sim.pc_utilization:
        ws_sum.cell(row=row, column=1, value="Utilización por PC (segundos)").font = section_font
        row += 1
        pc_headers = ["PC", "Ocupado (s)", "Mantenimiento (s)", "Libre (s)"]
        for i, h in enumerate(pc_headers, start=1):
            c = ws_sum.cell(row=row, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        row += 1
        for pc in json.loads(sim.pc_utilization):
            ws_sum.cell(row=row, column=1, value=f"PC {pc.get('id')}")
            ws_sum.cell(row=row, column=2, value=round(pc.get("busy_time", 0), 2))
            ws_sum.cell(row=row, column=3, value=round(pc.get("maintenance_time", 0), 2))
            ws_sum.cell(row=row, column=4, value=round(pc.get("idle_time", 0), 2))
            row += 1

    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 22
    ws_sum.column_dimensions["D"].width = 22

    # --- Sheet 2: Vector de Estados ---
    # Mantiene el vector completo con snapshots JSON para PCs, encargado y alumnos activos.
    ws_vec = wb.create_sheet("Vector de Estados")
    line_headers = [
        "Fila", "Reloj (formato)", "Reloj (seg)", "Evento", "Cola",
        "RND Llegada", "Tpo Llegada (seg)", "Próx. Llegada (seg)", "Alumnos Rechazados",
        "Estados PCs", "Snapshot PCs", "Encargado", "Cola IDs", "Detalle alumnos activos",
        "RND Inscripción", "Tpo Inscripción (seg)", "Inscripciones Completadas",
        "RND Mantenimiento", "Tpo Mantenimiento (seg)",
        "RND Regreso Técnico", "Tpo Regreso Técnico (seg)", "Próx. Mantenimiento (seg)", "Fin Mantenimiento (seg)",
    ]
    for i, h in enumerate(line_headers, start=1):
        c = ws_vec.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    q = db.query(SimulationLineModel)\
          .filter(SimulationLineModel.simulation_id == simulation_id)\
          .order_by(SimulationLineModel.line_index.asc())\
          .yield_per(1000)

    row = 2
    for ln in q:
        ws_vec.cell(row=row, column=1, value=ln.line_index)
        ws_vec.cell(row=row, column=2, value=ln.clock_formatted)
        ws_vec.cell(row=row, column=3, value=ln.clock)
        ws_vec.cell(row=row, column=4, value=ln.event_name)
        ws_vec.cell(row=row, column=5, value=ln.queue_length)
        ws_vec.cell(row=row, column=6, value=ln.student_rnd)
        ws_vec.cell(row=row, column=7, value=ln.student_arrival_time)
        ws_vec.cell(row=row, column=8, value=ln.student_next_arrival_time)
        ws_vec.cell(row=row, column=9, value=ln.total_students_returned)
        ws_vec.cell(row=row, column=10, value=ln.pc_states)
        ws_vec.cell(row=row, column=11, value=ln.pc_snapshot_json)
        ws_vec.cell(row=row, column=12, value=ln.encargado_snapshot_json)
        ws_vec.cell(row=row, column=13, value=ln.queue_student_ids_json)
        ws_vec.cell(row=row, column=14, value=ln.active_students_snapshot_json)
        ws_vec.cell(row=row, column=15, value=ln.registration_rnd)
        ws_vec.cell(row=row, column=16, value=ln.registration_time)
        ws_vec.cell(row=row, column=17, value=ln.registrations_completed)
        ws_vec.cell(row=row, column=18, value=ln.maintenance_rnd)
        ws_vec.cell(row=row, column=19, value=ln.maintenance_time)
        ws_vec.cell(row=row, column=20, value=ln.technician_return_rnd)
        ws_vec.cell(row=row, column=21, value=ln.technician_return_time)
        ws_vec.cell(row=row, column=22, value=ln.next_maintenance_start_time)
        ws_vec.cell(row=row, column=23, value=ln.next_maintenance_complete_time)
        row += 1

    widths = [8, 14, 12, 28, 8, 12, 14, 16, 14, 16, 36, 42, 18, 50, 14, 16, 16, 14, 16, 16, 18, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws_vec.column_dimensions[ws_vec.cell(row=1, column=i).column_letter].width = w
    ws_vec.freeze_panes = "A2"

    # --- Sheet 3: Detalle Alumnos ---
    # Auditoría histórica: también incluye alumnos destruidos por inscripción o rechazo final.
    ws_students = wb.create_sheet("Detalle Alumnos")
    student_headers = [
        "Alumno", "Estado final", "Intentos", "Veces que volvió", "Espera total (seg)",
        "Primera llegada (seg)", "Último evento (seg)", "Minuto de vuelta (seg)",
        "Inscripción completada (seg)",
    ]
    for i, h in enumerate(student_headers, start=1):
        c = ws_students.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    row = 2
    student_q = db.query(SimulationStudentModel)\
        .filter(SimulationStudentModel.simulation_id == simulation_id)\
        .order_by(SimulationStudentModel.student_id.asc())\
        .yield_per(1000)
    for student in student_q:
        ws_students.cell(row=row, column=1, value=student.student_id)
        ws_students.cell(row=row, column=2, value=student.final_state)
        ws_students.cell(row=row, column=3, value=student.attempts)
        ws_students.cell(row=row, column=4, value=student.times_returned_later)
        ws_students.cell(row=row, column=5, value=student.total_waiting_time)
        ws_students.cell(row=row, column=6, value=student.first_arrival_time)
        ws_students.cell(row=row, column=7, value=student.last_event_time)
        ws_students.cell(row=row, column=8, value=student.return_time)
        ws_students.cell(row=row, column=9, value=student.completed_registration_at)
        row += 1
    for i, w in enumerate([10, 18, 10, 16, 18, 20, 18, 18, 24], start=1):
        ws_students.column_dimensions[ws_students.cell(row=1, column=i).column_letter].width = w
    ws_students.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"simulacion_{sim.id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.delete("/api/simulations/{simulation_id}")
def delete_simulation(simulation_id: int, db: Session = Depends(get_db)):
    """
    Deletes a simulation run and all of its associated lines.
    """
    sim = db.query(SimulationModel).filter(SimulationModel.id == simulation_id).first()
    if not sim:
        raise HTTPException(status_code=404, detail="Simulation not found")
        
    db.delete(sim)
    db.commit()
    return {"message": f"Simulation {simulation_id} deleted successfully"}
