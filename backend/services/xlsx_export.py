import io
import json
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_workbook(sim, lines_iter: Iterable, students_iter: Iterable) -> io.BytesIO:
    """Build the 3-sheet .xlsx (Resumen, Vector de Estados, Detalle Alumnos).

    Pure: no DB and no HTTP. `lines_iter` / `students_iter` are iterables of ORM
    line / student rows (streamed by the repository).
    """
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    section_font = Font(bold=True, size=12, color="6366F1")

    # --- Sheet 1: Resumen ---
    ws_sum = wb.active
    ws_sum.title = "Resumen"
    ws_sum["A1"] = f"Simulación #{sim.id}"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Ejecutada: {sim.created_at}"

    sections = [
        ("Parámetros", [
            ("Cantidad de PCs", sim.num_pcs),
            ("Tiempo entre llegadas (min)", sim.mean_arrival_time),
            ("Inscripción mín (min)", sim.min_enrollment),
            ("Inscripción máx (min)", sim.max_enrollment),
            ("Mantenimiento mín (min)", sim.min_maintenance_time),
            ("Mantenimiento máx (min)", sim.max_maintenance_time),
            ("Frecuencia regreso técnico (min)", sim.mean_technician_return_time),
            ("Variación regreso técnico ± (min)", sim.technician_return_time_variation),
            ("Límite cola alumnos", sim.student_wait_threshold),
            ("Demora retorno alumno (min)", sim.student_return_time),
            ("Mantenimiento inicial en minuto 0", "Sí" if sim.initial_maintenance_at_start else "No"),
            ("Tiempo simulado (días)", sim.sim_days),
            ("Tiempo simulado (horas)", round(sim.sim_days * 24, 2)),
        ]),
        ("Métricas", [
            ("Alumnos arribados (totales)", sim.total_students_arrived),
            ("Alumnos nuevos", sim.total_new_students_arrived),
            ("Alumnos retirados (rechazados)", sim.total_students_returned),
            ("Inscripciones completadas", sim.registrations_completed),
            ("Visitas del técnico", sim.total_technician_visits),
            ("% intentos rechazados", round(sim.pct_students_returned, 4)),
            ("Espera promedio (seg)", round(sim.avg_waiting_time, 4)),
            ("Espera promedio (min)", round(sim.avg_waiting_time / 60.0, 4)),
            ("Ocio promedio técnico (seg)", round(sim.avg_technician_idle_time, 4)),
            ("Ocio promedio técnico (min)", round(sim.avg_technician_idle_time / 60.0, 4)),
        ]),
    ]

    row = 4
    for title, rows in sections:
        ws_sum.cell(row=row, column=1, value=title).font = section_font
        row += 1
        for label, val in rows:
            ws_sum.cell(row=row, column=1, value=label)
            ws_sum.cell(row=row, column=2, value=val)
            row += 1
        row += 1

    if sim.pc_utilization:
        ws_sum.cell(row=row, column=1, value="Utilización por PC (segundos)").font = section_font
        row += 1
        pc_headers = ["PC", "Ocupado (s)", "Mantenimiento (s)", "Libre (s)"]
        for i, h in enumerate(pc_headers, start=1):
            c = ws_sum.cell(row=row, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        row += 1
        for pc in json.loads(sim.pc_utilization):
            ws_sum.cell(row=row, column=1, value=f"PC {pc.get('id')}")
            ws_sum.cell(row=row, column=2, value=round(pc.get("busy_time", 0), 2))
            ws_sum.cell(row=row, column=3, value=round(pc.get("maintenance_time", 0), 2))
            ws_sum.cell(row=row, column=4, value=round(pc.get("idle_time", 0), 2))
            row += 1

    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 22
    ws_sum.column_dimensions["D"].width = 22

    # --- Sheet 2: Vector de Estados ---
    ws_vec = wb.create_sheet("Vector de Estados")
    line_headers = [
        "Fila", "Reloj (formato)", "Reloj (seg)", "Evento", "Cola",
        "RND Llegada", "Tpo Llegada (seg)", "Próx. Llegada (seg)", "Alumnos Rechazados",
        "Estados PCs", "Snapshot PCs", "Encargado", "Cola IDs", "Detalle alumnos activos",
        "RND Inscripción", "Tpo Inscripción (seg)", "Inscripciones Completadas",
        "RND Mantenimiento", "Tpo Mantenimiento (seg)",
        "RND Regreso Técnico", "Tpo Regreso Técnico (seg)", "Próx. Mantenimiento (seg)", "Fin Mantenimiento (seg)",
    ]
    for i, h in enumerate(line_headers, start=1):
        c = ws_vec.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    row = 2
    for ln in lines_iter:
        ws_vec.cell(row=row, column=1, value=ln.line_index)
        ws_vec.cell(row=row, column=2, value=ln.clock_formatted)
        ws_vec.cell(row=row, column=3, value=ln.clock)
        ws_vec.cell(row=row, column=4, value=ln.event_name)
        ws_vec.cell(row=row, column=5, value=ln.queue_length)
        ws_vec.cell(row=row, column=6, value=ln.student_rnd)
        ws_vec.cell(row=row, column=7, value=ln.student_arrival_time)
        ws_vec.cell(row=row, column=8, value=ln.student_next_arrival_time)
        ws_vec.cell(row=row, column=9, value=ln.total_students_returned)
        ws_vec.cell(row=row, column=10, value=ln.pc_states)
        ws_vec.cell(row=row, column=11, value=ln.pc_snapshot_json)
        ws_vec.cell(row=row, column=12, value=ln.encargado_snapshot_json)
        ws_vec.cell(row=row, column=13, value=ln.queue_student_ids_json)
        ws_vec.cell(row=row, column=14, value=ln.active_students_snapshot_json)
        ws_vec.cell(row=row, column=15, value=ln.registration_rnd)
        ws_vec.cell(row=row, column=16, value=ln.registration_time)
        ws_vec.cell(row=row, column=17, value=ln.registrations_completed)
        ws_vec.cell(row=row, column=18, value=ln.maintenance_rnd)
        ws_vec.cell(row=row, column=19, value=ln.maintenance_time)
        ws_vec.cell(row=row, column=20, value=ln.technician_return_rnd)
        ws_vec.cell(row=row, column=21, value=ln.technician_return_time)
        ws_vec.cell(row=row, column=22, value=ln.next_maintenance_start_time)
        ws_vec.cell(row=row, column=23, value=ln.next_maintenance_complete_time)
        row += 1

    widths = [8, 14, 12, 28, 8, 12, 14, 16, 14, 16, 36, 42, 18, 50, 14, 16, 16, 14, 16, 16, 18, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws_vec.column_dimensions[ws_vec.cell(row=1, column=i).column_letter].width = w
    ws_vec.freeze_panes = "A2"

    # --- Sheet 3: Detalle Alumnos ---
    ws_students = wb.create_sheet("Detalle Alumnos")
    student_headers = [
        "Alumno", "Estado final", "Intentos", "Veces que volvió", "Espera total (seg)",
        "Primera llegada (seg)", "Último evento (seg)", "Minuto de vuelta (seg)",
        "Inscripción completada (seg)",
    ]
    for i, h in enumerate(student_headers, start=1):
        c = ws_students.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    row = 2
    for student in students_iter:
        ws_students.cell(row=row, column=1, value=student.student_id)
        ws_students.cell(row=row, column=2, value=student.final_state)
        ws_students.cell(row=row, column=3, value=student.attempts)
        ws_students.cell(row=row, column=4, value=student.times_returned_later)
        ws_students.cell(row=row, column=5, value=student.total_waiting_time)
        ws_students.cell(row=row, column=6, value=student.first_arrival_time)
        ws_students.cell(row=row, column=7, value=student.last_event_time)
        ws_students.cell(row=row, column=8, value=student.return_time)
        ws_students.cell(row=row, column=9, value=student.completed_registration_at)
        row += 1
    for i, w in enumerate([10, 18, 10, 16, 18, 20, 18, 18, 24], start=1):
        ws_students.column_dimensions[ws_students.cell(row=1, column=i).column_letter].width = w
    ws_students.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
